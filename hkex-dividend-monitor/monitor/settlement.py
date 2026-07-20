"""Live settlement-price lookups from HKEX, SGX, and Eurex.

Three independent exchanges, three different (undocumented) data shapes --
each section below fetches and normalizes one of them into a plain
dict/list of dicts the web API and chat tools can serve directly. All
parsing is deterministic (regex/openpyxl/JSON navigation) -- no LLM is
involved anywhere in this module, unlike the dividend-filing pipeline.

  HKEX  -- a CMS-driven JSON endpoint behind the public Final Settlement
           Prices page; ~1 year of history for every listed contract in
           one call.
  SGX   -- a GraphQL-ish content API that resolves to a daily settlement
           price Excel workbook (openpyxl parses it).
  Eurex -- a JSON statistics API keyed by an internal numeric product id
           (not the public product code), plus a separate Excel workbook
           for MSCI futures final settlement prices.

Every fetch function raises SettlementError on failure; callers (web.py,
chat.py) catch it, log via monitor.diagnostics/monitor.activity, and
report a clean error rather than crashing. Each fetch function is wrapped
in a small in-process TTL cache (see _cached_fetch) so repeated tab loads
or chat turns within a few minutes don't re-hit these sites/re-parse the
same workbook; pass force=True to bypass it (used by dashboard Refresh
buttons).
"""
from __future__ import annotations

import json
import re
import threading
import time
from datetime import date as _date
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Optional

import openpyxl
import requests

from monitor.config import EUREX_PRODUCT_IDS_FILE, HKT
from monitor.jsonutil import atomic_write_json, load_json

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}


class SettlementError(RuntimeError):
    pass


# ---- Shared cache ----

_DEFAULT_TTL_SECONDS = 600.0
_CACHE: dict[str, tuple[float, Any]] = {}
_CACHE_LOCK = threading.Lock()
# A year-long sgx_daily walk stores ~250 sgx_daily:{key} entries (each
# ~0.5MB as JSON); left unbounded, a long-lived process's cache grows
# without limit. Capped and evicted oldest-stored-first -- generously
# sized (not e.g. 64) so a walk like that doesn't repeatedly evict and
# force-rebuild the much-more-expensive contract_cards/hkex_fsp entries.
_CACHE_MAX_ENTRIES = 128


def _cached_fetch(key: str, force: bool, ttl: float, fn):
    if not force:
        with _CACHE_LOCK:
            hit = _CACHE.get(key)
        if hit is not None and (time.monotonic() - hit[0]) < ttl:
            return hit[1]
    # Deliberately NOT held during fn() -- that's a network call, and
    # holding the lock across it would serialize every unrelated fetch in
    # the process behind whichever one happens to be in flight. Two
    # threads racing on the same cold key just both fetch once; the lock
    # only protects the dict itself, so the last write wins cleanly rather
    # than corrupting _CACHE's internal state.
    value = fn()
    with _CACHE_LOCK:
        _CACHE[key] = (time.monotonic(), value)
        while len(_CACHE) > _CACHE_MAX_ENTRIES:
            oldest_key = min(_CACHE, key=lambda k: _CACHE[k][0])
            if oldest_key == key:
                break  # never evict the entry this call just wrote
            del _CACHE[oldest_key]
    return value


def _cache_peek(key: str) -> Optional[tuple[float, Any]]:
    """Read a cache entry without affecting its TTL/eviction bookkeeping --
    (age_seconds, value) if present, else None. Lets a caller inspect what
    build_contract_cards last cached (e.g. whether it carried a
    sourcesFailed) without forcing a fetch of its own."""
    with _CACHE_LOCK:
        hit = _CACHE.get(key)
    if hit is None:
        return None
    stored_at, value = hit
    return time.monotonic() - stored_at, value


# ---- Shared helpers ----


def _download(url: str, timeout: float = 60.0) -> bytes:
    try:
        resp = requests.get(url, headers=_HEADERS, timeout=timeout)
        resp.raise_for_status()
        return resp.content
    except requests.RequestException as exc:
        raise SettlementError(f"Failed to download {url}: {exc}") from exc


def _dig(node: Any, *keys: str) -> Any:
    """Navigate a nested dict, raising SettlementError with a clear message
    on any missing key instead of a bare KeyError -- these shapes come
    from undocumented third-party APIs, so a missing key is an expected
    failure mode, not a bug."""
    for k in keys:
        if not isinstance(node, dict) or k not in node:
            raise SettlementError(f"Unexpected response shape (missing {k!r})")
        node = node[k]
    return node


def _load_workbook(raw_bytes: bytes, source: str) -> "openpyxl.Workbook":
    try:
        return openpyxl.load_workbook(BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any bad/corrupt workbook must not crash the caller
        raise SettlementError(f"Failed to parse {source} workbook: {exc}") from exc


def _cell_to_iso(value: Any) -> Any:
    """openpyxl returns date/datetime cells as Python objects; normalize to
    an ISO date string so results are plain JSON. Other cell types pass
    through unchanged (str, int, float, None)."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, _date):
        return value.isoformat()
    return value


_NUMERIC_STRING_RE = re.compile(r"^[+-]?(\d{1,3}(,\d{3})+|\d+)(\.\d+)?$")


def _coerce_numeric(value: Any) -> Any:
    """Settlement-price fields must sort and compare correctly, but the
    source cells arrive inconsistently typed: HKEX's are always strings
    (everything goes through _strip_tags) and SGX/Eurex xlsx cells are str
    or float depending on how that particular cell was formatted upstream.
    A numeric-looking string (plain, or US-style thousands-comma-grouped)
    becomes a float; anything that doesn't match -- including strings
    Python's own float() would happily accept but that aren't genuine
    settlement figures ("nan", "inf", "-inf") -- passes through unchanged,
    so a placeholder value never gets silently dropped, and a bare float()
    call never turns "nan" into a NaN that then serializes as invalid JSON.
    The regex also rejects European decimal-comma formatting ("1.234,56")
    outright rather than mis-coercing it (naive comma-stripping would have
    silently divided that value by 1000)."""
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if _NUMERIC_STRING_RE.match(stripped):
            return float(stripped.replace(",", ""))
        return value
    return value


_YEAR_MONTH_RE = re.compile(r"^(\d{4})-(\d{1,2})$")
_FULL_ISO_DATE_RE = re.compile(r"^(\d{4}-\d{2})-\d{1,2}$")


def _normalize_year_month(value: str) -> str:
    """Zero-pad a "YYYY-M" expiry/contract-month filter value to "YYYY-MM",
    and truncate a full "YYYY-MM-DD" to its "YYYY-MM" prefix.

    Both filter_hkex_rows' expiry_month and filter_sgx_rows' contract_month
    match via string prefix against an already-zero-padded ISO date
    ("2026-05-28"), so an LLM-supplied "2026-5" (a single-digit month is a
    natural thing to type) would otherwise silently match zero rows instead
    of the intended month. The full-date case matters because a value
    copied verbatim from one tool's row into another's filter can be a
    complete date -- e.g. the SGX current-snapshot's contractMonth is a
    date ("2026-07-01") while sgx_daily/history filters expect "YYYY-MM";
    without this, that round-trip silently matches nothing. Non-YYYY-M,
    non-full-date input passes through unchanged (e.g. HKEX's own "May-26"
    wording, matched on a different field entirely).
    """
    value = value.strip()
    m = _YEAR_MONTH_RE.match(value)
    if m:
        year, month = m.groups()
        return f"{year}-{int(month):02d}"
    m = _FULL_ISO_DATE_RE.match(value)
    if m:
        return m.group(1)
    return value


def _coerce_months_back(value: Any) -> int:
    """months_back arrives as an LLM tool-call argument -- tolerate a
    numeric string ("3") the same way an int is tolerated; anything else
    (unparseable, negative -- which would compute a future cutoff and
    silently match nothing) is treated as "no filter" rather than raising
    or silently emptying the result."""
    try:
        n = int(value)
    except (TypeError, ValueError):
        return 0
    return n if n > 0 else 0


# ============================================================
# HKEX -- Final Settlement Prices
# ============================================================

HKEX_FSP_PAGE_URL = "https://www.hkex.com.hk/Services/Trading/Derivatives/Final-Settlement-Prices?sc_lang=en"
HKEX_FSP_JSON_BASE = "https://www.hkex.com.hk/Market/Json/fSPTabAndTable"

# Fallback GUIDs if the page-scrape below ever fails -- these are stable
# CMS content-table identifiers (not secrets), current as of 2026-07.
# Discovering them from the live page first means we self-heal if HKEX
# ever regenerates them; this is just a safety net for a transient
# page-fetch hiccup.
_HKEX_FALLBACK_TABLE_ID = "D3EAB5EFE4C14F4CB96457F424B1A8BA"
_HKEX_FALLBACK_SORT_COL = "71AD4C35FDE4499A81BB93CE0A09DBE1"

_HKEX_TABLE_ATTR_RE = re.compile(
    r'<table[^>]*\bdata-table-id="([A-Fa-f0-9]+)"[^>]*\bdata-sort-col="([A-Fa-f0-9]+)"'
)
_HKEX_TAG_RE = re.compile(r"<[^>]+>")
_HKEX_DATE_RE = re.compile(r"^\d{1,2}-[A-Za-z]{3}-\d{4}$")
_HKEX_NON_DATA_KEYS = ("YearMonth", "RowKey", "RowPopupContent")


def _strip_tags(value: Any) -> str:
    if value is None:
        return ""
    return _HKEX_TAG_RE.sub("", str(value)).strip()


def _looks_like_number(value: str) -> bool:
    if not value:
        return False
    try:
        float(value.replace(",", ""))
        return True
    except ValueError:
        return False


def _looks_like_date(value: str) -> bool:
    return bool(_HKEX_DATE_RE.match(value))


def _parse_hkex_date(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%d-%b-%Y").date().isoformat()
    except ValueError:
        return None


def _discover_hkex_table_ids() -> tuple[str, str]:
    """Scrape the public FSP page for its `data-table-id`/`data-sort-col`
    attributes -- these drive the JSON endpoint URL and aren't published
    anywhere else. Falls back to the last-known-good pair on any failure
    (network error, or HKEX changing the page markup) rather than raising,
    since the JSON endpoint itself still works with a stale-but-valid pair
    until HKEX actually rotates them.
    """
    try:
        resp = requests.get(HKEX_FSP_PAGE_URL, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
        m = _HKEX_TABLE_ATTR_RE.search(resp.text)
        if m:
            return m.group(1), m.group(2)
    except requests.RequestException:
        pass
    return _HKEX_FALLBACK_TABLE_ID, _HKEX_FALLBACK_SORT_COL


def _fetch_hkex_json(table_id: str, sort_col: str) -> dict[str, Any]:
    # de = sort descending, English; _1 = page 1 (the table's own
    # maxNumOfFile has always been 1 in practice -- HKEX returns the whole
    # ~1-year history in a single JSON file, not paginated).
    url = f"{HKEX_FSP_JSON_BASE}/{table_id}_{sort_col}de_1.json"
    try:
        resp = requests.get(url, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
        return resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise SettlementError(f"HKEX settlement price fetch failed: {exc}") from exc


# How many rows the column-shape sniff in _map_hkex_columns samples. Plenty
# to out-vote a few odd cells, small enough to stay O(1)-ish on a ~2k-row
# payload.
_HKEX_SNIFF_SAMPLE_ROWS = 50


def _sniff_column_kind(rows: list[dict[str, Any]], guid: str) -> str:
    """Classify one unmapped HKEX column as publishDate / fsp / productType
    by majority vote over a sample of its non-empty values -- NOT just the
    first row's, where a single blank/placeholder cell (e.g. an "N/A"
    settlement price) would silently mislabel the whole column for every
    row in the table."""
    votes = {"publishDate": 0, "fsp": 0, "productType": 0}
    sampled = 0
    for row in rows:
        value = _strip_tags(row.get(guid))
        if not value:
            continue
        if _looks_like_date(value):
            votes["publishDate"] += 1
        elif _looks_like_number(value):
            votes["fsp"] += 1
        else:
            votes["productType"] += 1
        sampled += 1
        if sampled >= _HKEX_SNIFF_SAMPLE_ROWS:
            break
    if sampled == 0:
        # An entirely-empty column carries no signal; "productType" (the
        # text bucket, and the old single-row check's fallback for "")
        # is the safe label -- it must NOT default to publishDate/fsp,
        # where a second column mapping to the same field name would
        # overwrite the real column's values in _normalize_hkex_row.
        return "productType"
    # max() is stable on ties in dict-insertion order: date > number > text,
    # the same precedence the old single-row check applied.
    return max(votes, key=votes.get)


def _map_hkex_columns(payload: dict[str, Any], sort_col_guid: str) -> dict[str, str]:
    """Map HKEX's opaque GUID column keys to semantic field names.

    The three filterable columns are self-describing via `searchOptions`
    (its `filterOption` label). The `sort_col_guid` (already known from
    page discovery) is HKEX's own default-sort row id, not a data field --
    excluded rather than heuristically guessed. Everything else is sniffed
    from its values' shape across a sample of rows (see _sniff_column_kind):
    a "DD-Mon-YYYY" string is a publish date, a bare number is the
    settlement price, anything else is the product type -- the only three
    kinds of column HKEX has ever shown here.
    """
    mapping: dict[str, str] = {sort_col_guid: "sortKey"}

    for guid, opt in (payload.get("searchOptions") or {}).items():
        filter_option = (opt or {}).get("filterOption", "") or ""
        if filter_option == "Contract":
            mapping[guid] = "contract"
        elif filter_option == "HKATS Code":
            mapping[guid] = "hkatsCode"
        elif "Trading Date" in filter_option or "Expiry" in filter_option:
            mapping[guid] = "lastTradingDate"

    rows = payload.get("tableInfo") or []
    if rows:
        for guid in rows[0].keys():
            if guid in mapping or guid in _HKEX_NON_DATA_KEYS:
                continue
            mapping[guid] = _sniff_column_kind(rows, guid)
    return mapping


def _normalize_hkex_row(raw: dict[str, Any], column_map: dict[str, str]) -> dict[str, Any]:
    row: dict[str, Any] = {}
    for guid, field in column_map.items():
        if field == "sortKey":
            continue
        row[field] = _strip_tags(raw.get(guid))
    if "fsp" in row:
        row["fsp"] = _coerce_numeric(row["fsp"])
    row["yearMonth"] = _strip_tags(raw.get("YearMonth"))
    row["publishDateIso"] = _parse_hkex_date(row.get("publishDate"))
    row["lastTradingDateIso"] = _parse_hkex_date(row.get("lastTradingDate"))
    return row


def _parse_hkex_gen_date(value: Any) -> Optional[str]:
    """HKEX's payload carries its own generation timestamp (`genDate`,
    live-observed as epoch seconds; treated as milliseconds instead if the
    raw value is too large to be seconds this century) -- surfaced as
    dataGeneratedAt so a caller can report when HKEX itself produced this
    data, distinct from asOf (when this app fetched/cached it -- can lag
    genDate by the cache TTL, or more if HKEX paused updating mid-day)."""
    if not isinstance(value, (int, float)) or value <= 0:
        return None
    epoch_seconds = value / 1000 if value > 1e12 else value
    try:
        return datetime.fromtimestamp(epoch_seconds, tz=HKT).isoformat()
    except (ValueError, OSError, OverflowError):
        return None


def _fetch_hkex_fsp_impl() -> dict[str, Any]:
    table_id, sort_col = _discover_hkex_table_ids()
    payload = _fetch_hkex_json(table_id, sort_col)
    if "tableInfo" not in payload:
        # Distinct from a present-but-empty list (which would mean HKEX
        # genuinely has zero rows today -- implausible, but a real possible
        # state) -- a MISSING key means the payload's shape itself changed
        # (a maintenance page, a redesigned endpoint), which silently
        # parsing as zero rows would hide as "no settlement prices exist"
        # rather than surfacing as the fetch problem it actually is.
        raise SettlementError("HKEX FSP payload is missing the expected 'tableInfo' key")
    column_map = _map_hkex_columns(payload, sort_col)
    rows = [_normalize_hkex_row(r, column_map) for r in payload.get("tableInfo") or []]
    return {
        "asOf": datetime.now(HKT).isoformat(),
        "dataGeneratedAt": _parse_hkex_gen_date(payload.get("genDate")),
        "rows": rows,
        "contracts": sorted({r["contract"] for r in rows if r.get("contract")}),
        "productTypes": sorted({r["productType"] for r in rows if r.get("productType")}),
    }


def fetch_hkex_fsp(force: bool = False) -> dict[str, Any]:
    """The full HKEX Final Settlement Price table (~1 year of history,
    every listed futures/options contract). Cached in-process; pass
    force=True (e.g. a dashboard Refresh click) to bypass."""
    return _cached_fetch("hkex_fsp", force, _DEFAULT_TTL_SECONDS, _fetch_hkex_fsp_impl)


# Common user-facing index abbreviations -> the wording HKEX actually uses
# in FSP contract names. HKEX never writes "HSCEI"/"HSTECH" in the main
# monthly contract rows (only in tangential ones like "HSCEI Dividend Point
# Index Futures"), so without this expansion the natural search term
# returns the wrong contracts and misses the one the user meant.
_HKEX_INDEX_ALIASES = {
    "HSI": "Hang Seng Index",
    "HSCEI": "Hang Seng China Enterprises Index",
    "HSTECH": "Hang Seng TECH Index",
}

# The reverse direction: a user is at least as likely to type the full
# wording ("hang seng", "hang seng index") as the raw abbreviation --
# live-confirmed a bare "hang seng" query top-ranked Hang Seng BANK (a
# real, unrelated HKEX-listed stock futures contract) over the Hang Seng
# INDEX, since the index's own official name is longer and so scores
# lower on precision. Matched longest-phrase-first against a query (see
# monitor.settlement_search.search_contracts) so "hang seng china
# enterprises" isn't shadowed by the bare "hang seng" entry checked after
# it. A mapped value of None is a deliberate BLOCK, not a miss: "Hang Seng
# Bank"/"Hang Seng Biotech..." are themselves real, unrelated HKEX/SGX-
# listed entities that also start with "hang seng" -- without blocking
# them, a query plainly about one of those would still trigger the INDEX
# alias on top of its own real name/code signal.
_HKEX_REVERSE_PHRASE_ALIASES: tuple[tuple[str, Optional[str]], ...] = (
    ("hang seng china enterprises index", "HSCEI"),
    ("hang seng china enterprises", "HSCEI"),
    ("hang seng tech index", "HSTECH"),
    ("hang seng tech", "HSTECH"),
    ("hang seng bank", None),
    ("hang seng biotech", None),
    ("hang seng index", "HSI"),
    ("hang seng", "HSI"),
)


def _hkats_components(row: dict[str, Any]) -> list[str]:
    """HKEX combines related contracts into one FSP row with a compound
    HKATS code like "HSI / MHI" or "HHI/MCH" -- match against each part,
    not the literal joined string."""
    return [p.strip().upper() for p in (row.get("hkatsCode") or "").split("/") if p.strip()]


def filter_hkex_rows(
    rows: list[dict[str, Any]],
    contract: Optional[str] = None,
    hkats_code: Optional[str] = None,
    months_back: Optional[int] = None,
    expiry_month: Optional[str] = None,
) -> list[dict[str, Any]]:
    """Narrow a fetch_hkex_fsp() row list -- shared by the chat tool (which
    needs a small result to reason over) and available to the API for the
    same reason. The dashboard tab itself filters client-side instead
    (same convention as the Dividend Watchlist table), since the full
    table is only ~1-2k rows.

    Contract matching unions two precise checks -- substring of the
    official contract name (with well-known index abbreviations like
    "HSCEI" also expanded to HKEX's own wording) and the query as an exact
    HKATS code component -- falling back to all-words-present for
    multi-word queries HKEX phrases differently. The code check must be a
    union, not a fallback: a short code can accidentally substring-match an
    unrelated name (e.g. "TCH" is inside "CK Hutchison"), which would
    otherwise silently shadow the contract actually coded TCH.
    """
    out = rows
    if contract:
        # str() coercions here and below: these values arrive as LLM tool-call
        # arguments, which can be mistyped (an int, a number-like value) --
        # a wrong type should degrade to a text match, not an AttributeError.
        contract = str(contract)
        needle = contract.strip().lower()
        needles = {needle}
        alias = _HKEX_INDEX_ALIASES.get(contract.strip().upper())
        if alias:
            needles.add(alias.lower())
        as_code = contract.strip().upper()
        matched = [
            r
            for r in out
            if any(n in (r.get("contract") or "").lower() for n in needles)
            or as_code in _hkats_components(r)
        ]
        if not matched:
            tokens = needle.split()
            matched = [r for r in out if all(t in (r.get("contract") or "").lower() for t in tokens)]
        out = matched
    if hkats_code:
        needle = str(hkats_code).strip().upper()
        out = [r for r in out if needle in _hkats_components(r)]
    if expiry_month:
        # "Which expiry?" must be answered by this filter, not by the model
        # scanning dozens of near-identical rows for the right month (it has
        # picked a neighboring expiry and mislabeled its date before).
        # Accepts "YYYY-MM" (matched on last trading date) or HKEX's own
        # yearMonth wording ("May-26"). _normalize_year_month zero-pads a
        # single-digit month ("2026-5") before the ISO-shape check below.
        wanted = _normalize_year_month(str(expiry_month))
        if re.fullmatch(r"\d{4}-\d{2}", wanted):
            out = [r for r in out if (r.get("lastTradingDateIso") or "").startswith(wanted)]
        else:
            out = [r for r in out if (r.get("yearMonth") or "").lower() == wanted.lower()]
    months_back = _coerce_months_back(months_back)
    if months_back:
        cutoff = (datetime.now(HKT) - timedelta(days=31 * months_back)).date().isoformat()
        out = [r for r in out if (r.get("publishDateIso") or "") >= cutoff]
    return out


# ============================================================
# SGX -- Final Settlement Prices
# ============================================================

SGX_APPCONFIG_URL = "https://www.sgx.com/config/appconfig.json"
SGX_CMS_PAGE_PATH = "/derivatives/clearing-information"

_SGX_SHEET_FIELDS = ("productType", "contract", "ticker", "contractMonth", "fsp", "fspDate")
_SGX_SHEET_NAMES = ("Financials Contracts", "Commodities Contracts")


_SGX_APPCONFIG_TTL_SECONDS = 600.0


def _fetch_sgx_appconfig_impl() -> dict[str, Any]:
    try:
        resp = requests.get(SGX_APPCONFIG_URL, headers=_HEADERS, timeout=20)
        resp.raise_for_status()
        return resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise SettlementError(f"SGX site config fetch failed: {exc}") from exc


def _fetch_sgx_appconfig() -> dict[str, Any]:
    """Cached: a cold fetch_sgx_fsp() + fetch_sgx_flexc() call each resolve
    their own CMS endpoints independently, which used to mean this file was
    downloaded twice for one settlement-price question -- appconfig barely
    changes, so a short TTL cache shares one fetch between them."""
    return _cached_fetch("sgx_appconfig", False, _SGX_APPCONFIG_TTL_SECONDS, _fetch_sgx_appconfig_impl)


def _sgx_cms_query(cms_api_url: str, cms_version: str, query_id: str, variables: dict[str, Any]) -> dict[str, Any]:
    try:
        resp = requests.get(
            cms_api_url.rstrip("/") + "/",
            params={"queryId": f"{cms_version}:{query_id}", "variables": json.dumps(variables)},
            headers=_HEADERS,
            timeout=30,
        )
        resp.raise_for_status()
        payload = resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise SettlementError(f"SGX content query {query_id!r} failed: {exc}") from exc
    if isinstance(payload, dict) and payload.get("errors"):
        raise SettlementError(f"SGX content query {query_id!r} returned an error: {payload['errors']}")
    return payload


def _find_widget(node: Any, widget_type: str) -> Optional[dict[str, Any]]:
    """Depth-first search for the first widget dict with this widgetType,
    anywhere in an SGX CMS page response -- the page's widget list is
    nested several "data" wrappers deep and its exact shape isn't
    contractually stable, so search rather than hard-code a path."""
    if isinstance(node, dict):
        if node.get("widgetType") == widget_type:
            return node
        for v in node.values():
            found = _find_widget(v, widget_type)
            if found is not None:
                return found
    elif isinstance(node, list):
        for v in node:
            found = _find_widget(v, widget_type)
            if found is not None:
                return found
    return None


def _sgx_cms_endpoints() -> tuple[str, str]:
    appconfig = _fetch_sgx_appconfig()
    return _dig(appconfig, "endpoints", "CMS_API_URL"), _dig(appconfig, "CMS_VERSION")


def _sgx_main_file_url(cms_api_url: str, cms_version: str) -> str:
    page = _sgx_cms_query(cms_api_url, cms_version, "page", {"path": SGX_CMS_PAGE_PATH, "lang": "EN"})
    widget = _find_widget(page, "final_settlement_price")
    if not widget:
        raise SettlementError("SGX clearing-information page has no final_settlement_price widget")
    items = widget.get("downloadItems") or []
    if not items:
        raise SettlementError("SGX final_settlement_price widget has no downloadItems")
    return _dig(items[0], "data", "file", "data", "file", "data", "url")


_SGX_DDMMYY_RE = re.compile(r"^(\d{2})(\d{2})(\d{2})$")
_ISO_DATE_SHAPE_RE = re.compile(r"^(\d{4})-\d{2}-\d{2}$")


def _normalize_sgx_contract_month(value: Any) -> Any:
    """SGX's contractMonth column carries three different shapes depending
    on row type, all live-observed: a real date/datetime cell (already ISO
    via _cell_to_iso) for ordinary monthly contracts; a bare DDMMYY text
    string ("140726" = 14 Jul 2026) for weekly options, which this app's
    contract_month filtering (an ISO "YYYY-MM" prefix match) can never
    match as a month, leaving weekly rows permanently unreachable by month
    filter; and occasional Excel-serial-date garbage in SGX's own workbook
    (an ISO-shaped date with an implausible year like "1900-01-02" --
    Excel's day-zero, misread as a real date). DDMMYY becomes its "YYYY-MM"
    month (the week itself stays distinguishable via the contract's own
    name); an implausible-year date becomes None rather than a
    filterable-but-wrong month. A legitimate ISO date passes through
    unchanged -- this only strips out the one bad-year case."""
    if isinstance(value, str):
        stripped = value.strip()
        m = _SGX_DDMMYY_RE.match(stripped)
        if m:
            dd, mm, yy = m.groups()
            return f"20{yy}-{mm}"
        m = _ISO_DATE_SHAPE_RE.match(stripped)
        if m and int(m.group(1)) < 2000:
            return None
    return value


def _parse_sgx_workbook(raw_bytes: bytes) -> list[dict[str, Any]]:
    wb = _load_workbook(raw_bytes, "SGX settlement price")
    rows: list[dict[str, Any]] = []
    for sheet_name in _SGX_SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            continue
        row_iter = wb[sheet_name].iter_rows(values_only=True)
        next(row_iter, None)  # header row
        for raw_row in row_iter:
            if not raw_row or raw_row[0] is None:
                continue
            entry: dict[str, Any] = {"sheet": sheet_name}
            for field, value in zip(_SGX_SHEET_FIELDS, raw_row):
                value = _cell_to_iso(value)
                if field == "fsp":
                    value = _coerce_numeric(value)
                elif field == "contractMonth":
                    value = _normalize_sgx_contract_month(value)
                entry[field] = value
            rows.append(entry)
    return rows


def _fetch_sgx_fsp_impl() -> dict[str, Any]:
    cms_api_url, cms_version = _sgx_cms_endpoints()
    file_url = _sgx_main_file_url(cms_api_url, cms_version)
    rows = _parse_sgx_workbook(_download(file_url))
    return {"asOf": datetime.now(HKT).isoformat(), "sourceFileUrl": file_url, "rows": rows}


def fetch_sgx_fsp(force: bool = False) -> dict[str, Any]:
    """Today's SGX-DC Final Settlement Price workbook (Financials +
    Commodities Contracts sheets), downloaded live and parsed with
    openpyxl. Cached in-process; pass force=True to bypass."""
    return _cached_fetch("sgx_fsp", force, _DEFAULT_TTL_SECONDS, _fetch_sgx_fsp_impl)


def _sgx_ticker_components(ticker: str) -> list[str]:
    """SGX combines related tickers into one compound field like "NK/NKO"
    -- same pattern as HKEX's compound HKATS codes (see _hkats_components
    above) -- split so an exact-code lookup on just "NK" still finds a row
    whose own ticker field is "NK/NKO"."""
    return [p.strip().upper() for p in (ticker or "").split("/") if p.strip()]


# A search this short is a plausible ticker/code, but as a raw substring
# it can also land inside an unrelated word -- live-confirmed: "NK" is a
# substring of "Bank", so searching "NK" for Nikkei pulled in every SGX
# NIFTY Bank Index row alongside the real Nikkei ones. Below this length,
# require an exact ticker-component match or a whole-word match in the
# contract name instead of a bare substring; longer needles (multi-word
# phrases like "usd nikkei") keep the more permissive substring match.
_SGX_SHORT_SEARCH_LEN = 3


def filter_sgx_rows(
    rows: list[dict[str, Any]],
    search: Optional[str] = None,
    contract_month: Optional[str] = None,
) -> list[dict[str, Any]]:
    """Narrow a fetch_sgx_fsp() row list. One contract can have rows across
    several contract months (the same expiry-scanning risk HKEX's rows have
    -- see filter_hkex_rows), so a specific-expiry question should use
    contract_month rather than leaving the model to pick the right row out
    of a longer list."""
    out = rows
    if search:
        needle = str(search).strip()
        if len(needle) <= _SGX_SHORT_SEARCH_LEN:
            needle_upper = needle.upper()
            word_pattern = re.compile(rf"\b{re.escape(needle)}\b", re.IGNORECASE)
            out = [
                r for r in out
                if needle_upper in _sgx_ticker_components(r.get("ticker") or "")
                or word_pattern.search(r.get("contract") or "")
            ]
        else:
            needle_lower = needle.lower()
            out = [
                r for r in out
                if needle_lower in (r.get("contract") or "").lower()
                or needle_lower in (r.get("ticker") or "").lower()
            ]
    if contract_month:
        # _normalize_year_month zero-pads a single-digit month ("2026-7")
        # so it still prefix-matches contractMonth's zero-padded ISO form.
        wanted = _normalize_year_month(str(contract_month)).strip()
        out = [r for r in out if (r.get("contractMonth") or "").startswith(wanted)]
    return out


_FLEXC_MMDDYYYY_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})$")


def _normalize_flexc_date(value: Any) -> Any:
    """FlexC's fspDate cells are sometimes real date/datetime objects
    (already normalized to ISO by _cell_to_iso) and sometimes plain text --
    observed live, US-style MM/DD/YYYY (e.g. "07/01/2026" for 1 July 2026).
    Left as text, that value flows verbatim into the app's history archive
    (breaking its date-equality lookups and its ORDER BY fspDate string
    sort across year/month boundaries) and sits next to ISO dates from
    every other tool in the same chat result, inviting a DD/MM misread.
    Normalized to ISO here, at the parse boundary, same as every other
    date in this module; non-matching strings (already-ISO, or anything
    else) pass through unchanged."""
    if isinstance(value, str):
        m = _FLEXC_MMDDYYYY_RE.match(value.strip())
        if m:
            month, day, year = (int(g) for g in m.groups())
            try:
                return _date(year, month, day).isoformat()
            except ValueError:
                return value
    return value


def _parse_sgx_flexc_workbook(raw_bytes: bytes) -> list[dict[str, Any]]:
    wb = _load_workbook(raw_bytes, "SGX FlexC settlement price")
    sheet_name = "Final Settlement" if "Final Settlement" in wb.sheetnames else wb.sheetnames[0]
    rows: list[dict[str, Any]] = []
    row_iter = wb[sheet_name].iter_rows(values_only=True)
    next(row_iter, None)  # header row
    for raw_row in row_iter:
        if not raw_row or raw_row[0] is None:
            continue
        ticker, fsp, fsp_date = (list(raw_row) + [None, None, None])[:3]
        rows.append(
            {
                "ticker": ticker,
                "fsp": _coerce_numeric(_cell_to_iso(fsp)),
                "fspDate": _normalize_flexc_date(_cell_to_iso(fsp_date)),
            }
        )
    return rows


def _fetch_sgx_flexc_impl() -> dict[str, Any]:
    cms_api_url, cms_version = _sgx_cms_endpoints()
    payload = _sgx_cms_query(
        cms_api_url, cms_version, "fsp_files", {"fspFileType": "fsp_flexc", "lang": "EN", "limit": 1}
    )
    results = _dig(payload, "data", "list", "results")
    if not results:
        return {"asOf": datetime.now(HKT).isoformat(), "sourceFileUrl": None, "rows": []}
    file_url = _dig(results[0], "data", "file", "data", "url")
    rows = _parse_sgx_flexc_workbook(_download(file_url))
    return {"asOf": datetime.now(HKT).isoformat(), "sourceFileUrl": file_url, "rows": rows}


def fetch_sgx_flexc(force: bool = False) -> dict[str, Any]:
    """The most recent SGX-DC FlexC (flexible FX) Final Settlement Price
    file -- a narrower, separate workbook from fetch_sgx_fsp's main one."""
    return _cached_fetch("sgx_flexc", force, _DEFAULT_TTL_SECONDS, _fetch_sgx_flexc_impl)


# ============================================================
# Eurex -- daily settlement prices (exchange-traded)
# ============================================================

EUREX_STATS_URL_TMPL = "https://www.eurex.com/api/v1/overallstatistics/{product_id}"
EUREX_PRODUCTS_URL = "https://www.eurex.com/ex-en!dynSearch"

# Seed map of well-known product codes -> Eurex's internal numeric
# product id (the id the statistics API actually takes -- it is NOT the
# public product code and isn't listed in the product catalog itself).
# Each was resolved by fetching that product's page once and reading its
# embedded `data-product` attribute; see resolve_eurex_product_id_from_url
# for how a user adds any product not seeded here. Verified 2026-07.
_EUREX_SEED_PRODUCT_IDS: dict[str, int] = {
    "FDAX": 34642,  # DAX Futures
    "FESX": 34652,  # EURO STOXX 50 Index Futures
    "FBEU": 4663138,  # Euro-EU Bond Futures
}

_DATA_PRODUCT_RE = re.compile(r'data-product="(\d+)"')


def _load_resolved_eurex_ids() -> dict[str, int]:
    return load_json(EUREX_PRODUCT_IDS_FILE, {})


def _save_resolved_eurex_id(code: str, product_id: int) -> None:
    data = _load_resolved_eurex_ids()
    data[code.strip().upper()] = product_id
    atomic_write_json(EUREX_PRODUCT_IDS_FILE, data)


def resolve_eurex_product_id(code: str) -> Optional[int]:
    """Look up a Eurex numeric product id for `code` (e.g. "FDAX") from the
    seed map or the runtime-persisted store (data/eurex_product_ids.json).
    Returns None if unresolved -- callers should prompt for that product's
    Eurex page URL and call resolve_eurex_product_id_from_url instead."""
    code = code.strip().upper()
    if code in _EUREX_SEED_PRODUCT_IDS:
        return _EUREX_SEED_PRODUCT_IDS[code]
    return _load_resolved_eurex_ids().get(code)


def resolve_eurex_product_id_from_url(code: str, page_url: str) -> int:
    """Fetch a Eurex product page and extract its internal numeric product
    id from the embedded `data-product` attribute, then persist it under
    `code` (normalized upper-case) for future lookups. Used the one time a
    product isn't already in _EUREX_SEED_PRODUCT_IDS or the persisted
    store -- the user pastes the product's Eurex page URL once."""
    if not re.match(r"^https://www\.eurex\.com/", page_url):
        raise SettlementError("Eurex product page URL must be on https://www.eurex.com/")
    try:
        resp = requests.get(page_url, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise SettlementError(f"Failed to fetch Eurex product page: {exc}") from exc

    m = _DATA_PRODUCT_RE.search(resp.text)
    if not m:
        raise SettlementError("Could not find a product id on that Eurex page")
    product_id = int(m.group(1))
    _save_resolved_eurex_id(code, product_id)
    return product_id


def _fetch_eurex_products_impl() -> list[dict[str, Any]]:
    try:
        resp = requests.get(EUREX_PRODUCTS_URL, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
        payload = resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise SettlementError(f"Eurex product catalog fetch failed: {exc}") from exc
    if not payload.get("items"):
        # A missing/empty catalog silently degrades to "0 Eurex products
        # exist" -- implausible (this catalog has ~3000 entries) and, via
        # settlement_search's contract-card index, blanks every Eurex card
        # for up to the full cards-cache TTL. Surface as a fetch problem
        # instead of a valid-but-empty result.
        raise SettlementError("Eurex product catalog payload has no items")
    return [
        {
            "code": item.get("PRODUCT_ID"),
            "name": item.get("PRODUCT_NAME"),
            "group": item.get("PRODUCT_GROUP"),
            "currency": item.get("CURRENCY"),
        }
        for item in payload.get("items") or []
        if item.get("PRODUCT_ID")
    ]


_EUREX_PRODUCTS_TTL_SECONDS = 24 * 60 * 60  # ~3MB catalog that barely changes day to day


def fetch_eurex_products(force: bool = False) -> list[dict[str, Any]]:
    """The full Eurex product catalog (code/name/group/currency, ~3000
    entries) -- feeds the dashboard's product picker. Does not include
    which products already have a resolved numeric id; cross-reference
    with resolve_eurex_product_id per code."""
    return _cached_fetch("eurex_products", force, _EUREX_PRODUCTS_TTL_SECONDS, _fetch_eurex_products_impl)


def _parse_eurex_contract_date(value: Optional[str]) -> Optional[str]:
    if not value or len(value) != 8:
        return None
    try:
        return datetime.strptime(value, "%Y%m%d").date().isoformat()
    except ValueError:
        return None


def _fetch_eurex_settlement_impl(product_id: int, busdate: Optional[str]) -> dict[str, Any]:
    params = {"filtertype": "overview"}
    if busdate:
        params["busdate"] = busdate
    try:
        resp = requests.get(
            EUREX_STATS_URL_TMPL.format(product_id=product_id), params=params, headers=_HEADERS, timeout=30
        )
        resp.raise_for_status()
        payload = resp.json()
    except (requests.RequestException, ValueError) as exc:
        raise SettlementError(f"Eurex settlement price fetch failed: {exc}") from exc
    if isinstance(payload, dict) and "error" in payload:
        detail = payload["error"]
        raise SettlementError(f"Eurex API error: {detail.get('message', detail) if isinstance(detail, dict) else detail}")

    header = payload.get("header") or {}
    meta = payload.get("meta") or {}
    rows = [
        {
            "date": r.get("date"),
            "dateIso": _parse_eurex_contract_date(r.get("date")),
            "open": r.get("open"),
            "high": r.get("high"),
            "low": r.get("low"),
            "last": r.get("last"),
            "settlementPrice": r.get("dSettle"),
            "volume": r.get("volume"),
            "openInterest": r.get("openInt"),
            "contractType": r.get("contractType"),
        }
        for r in payload.get("dataRows") or []
    ]
    return {
        "asOf": datetime.now(HKT).isoformat(),
        "productId": product_id,
        "productCode": meta.get("productCode"),
        "isin": meta.get("isin"),
        "underlyingClosingPrice": header.get("underlyingClosingPrice"),
        "tradingDates": header.get("tradingDates") or [],
        "rows": rows,
    }


def fetch_eurex_settlement(product_id: int, busdate: Optional[str] = None, force: bool = False) -> dict[str, Any]:
    """Daily settlement prices (dSettle) per contract month for one Eurex
    product, as of `busdate` (Eurex's own "YYYYMMDD" format; defaults to
    its latest business date). Resolve `product_id` via
    resolve_eurex_product_id(code) first."""
    key = f"eurex_settlement:{product_id}:{busdate or ''}"
    return _cached_fetch(key, force, _DEFAULT_TTL_SECONDS, lambda: _fetch_eurex_settlement_impl(product_id, busdate))


# ============================================================
# Eurex -- MSCI futures final settlement prices (Excel)
# ============================================================

EUREX_MSCI_PAGE_URL = "https://www.eurex.com/ex-en/markets/idx/msci"
_MSCI_BLOB_RE = re.compile(r"(/resource/blob/\d+/[0-9a-f]+/data/msci-fut-settlement-prices\.xlsx)", re.IGNORECASE)
_MSCI_FIXED_COLUMNS = (
    "indexName",
    "region",
    "indexType",
    "markets",
    "currency",
    "dividendReinvestment",
    "eurexCode",
    "futuresBbg",
)


def _parse_msci_workbook(raw_bytes: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = _load_workbook(raw_bytes, "Eurex MSCI settlement price")
    sheet_name = "FSP MSCI Futures" if "FSP MSCI Futures" in wb.sheetnames else wb.sheetnames[0]
    row_iter = wb[sheet_name].iter_rows(values_only=True)
    next(row_iter, None)  # title row
    header_row = next(row_iter, None) or ()
    expiries = [str(h).strip() for h in header_row[len(_MSCI_FIXED_COLUMNS) :] if h]

    rows: list[dict[str, Any]] = []
    for raw_row in row_iter:
        if not raw_row or not raw_row[0]:
            continue
        entry: dict[str, Any] = dict(zip(_MSCI_FIXED_COLUMNS, raw_row))
        by_expiry = {
            expiry: _coerce_numeric(value)
            for expiry, value in zip(expiries, raw_row[len(_MSCI_FIXED_COLUMNS) :])
            if value is not None
        }
        if not entry.get("eurexCode") and not by_expiry:
            # A legend/footnote row (e.g. "* DM = Developed Markets / EM =
            # Emerging Markets / FM = Frontier Markets Futures") has text
            # in its first cell -- passing the blank-leading-cell skip
            # above -- but no real product code and no settlement figures
            # at all. Left in, it becomes a retrievable "contract" with a
            # null price.
            continue
        entry["settlementPricesByExpiry"] = by_expiry
        rows.append(entry)
    return rows, expiries


def _fetch_eurex_msci_fsp_impl() -> dict[str, Any]:
    try:
        resp = requests.get(EUREX_MSCI_PAGE_URL, headers=_HEADERS, timeout=30)
        resp.raise_for_status()
    except requests.RequestException as exc:
        raise SettlementError(f"Eurex MSCI page fetch failed: {exc}") from exc

    m = _MSCI_BLOB_RE.search(resp.text)
    if not m:
        raise SettlementError("Could not find the MSCI settlement price file link on the Eurex MSCI page")
    file_url = "https://www.eurex.com" + m.group(1)

    rows, expiries = _parse_msci_workbook(_download(file_url))
    return {
        "asOf": datetime.now(HKT).isoformat(),
        "sourceFileUrl": file_url,
        "expiries": expiries,
        "rows": rows,
    }


def fetch_eurex_msci_fsp(force: bool = False) -> dict[str, Any]:
    """Eurex MSCI Futures final settlement prices -- one row per MSCI
    index, with a settlementPricesByExpiry dict keyed by expiry column
    (e.g. "FSP MAR18"). Use latest_populated_msci_expiry to pick a
    sensible default column to display."""
    return _cached_fetch("eurex_msci_fsp", force, _DEFAULT_TTL_SECONDS, _fetch_eurex_msci_fsp_impl)


def latest_populated_msci_expiry(rows: list[dict[str, Any]], expiries: list[str]) -> Optional[str]:
    """The rightmost expiry column with at least one non-empty value across
    all rows -- the dashboard's default view (expiry columns run oldest to
    newest left to right, and older ones are all blank once superseded)."""
    for expiry in reversed(expiries):
        if any(r.get("settlementPricesByExpiry", {}).get(expiry) is not None for r in rows):
            return expiry
    return None
